"""
封装一个读取用例的excel类
# 实现读取用例数据
# 实现写入数据的功能
"""
from collections import namedtuple

import openpyxl


class Case:
    # 这个类用来存储用例
    pass


class ReadExcel(object):
    """
    读取excel数据
    """

    def __init__(self, file_name, sheet_name):
        """
        这个是用例初始化读取对象的
        :param file_name:  文件名字  -->  str
        :param sheet_name: 表单名字  -->  str
        """
        self.filename = file_name
        # 打开工作簿
        self.wb = openpyxl.load_workbook(file_name)
        # 选择表单
        self.sheet = self.wb[sheet_name]

    def __del__(self):
        # self.wb.save(self.filename)
        self.wb.close()

    def r_data_obj(self):
        """
        按行读取数据，表单所有数据
        每个用例存储在一个对象中
        :return: 返回一个列表，列表中每个元素为一个用例对象
        """
        # 按行获取数据转换成列表
        rows_data = list(self.sheet.rows)
        # 获取表单的表头信息
        titles = []
        for title in rows_data[0]:
            titles.append(title.value)
        # 定义一个空列表用来存储所有的用例
        cases = []
        for case in rows_data[1:]:
            # 创建一个Cases类的对象，用来保存用例数据，
            case_obj = Case()
            # data用例临时存放用例数据
            data = []
            # 判断该单元格是否为字符串类型，
            for cell in case:
                data.append(cell.value)
            # 将该条数据放入cases中
            case_data = list(zip(titles, data))
            for i in case_data:
                setattr(case_obj, i[0], i[1])
            cases.append(case_obj)

        return cases

    def r_data_obj_from_line(self, list1):
        """
        按list中行读取，表单所有该行数据
        每个用例存储在一个对象中
        :param list1: list1  -->   要求读取的行编号列表，eg:[1,3,5]则会读取第1,3,5条用例
        :return: 包含所有用例对象的列表list
        """
        # list1 参数为一个列表，传入的是指定读取数据的列,比如[1,2,3]
        # 每一行[1,3,5]列的数据，读取出来就作为一条测试用例，放在对象中属性中
        # 所有的用例对象放在列表中并且进行返回
        rows_data = list(self.sheet.rows)
        titles = []
        for title in rows_data[0]:
            titles.append(title.value)
        case_all = []
        for k in list1:
            Cases = namedtuple('Cases', ['%s' % x for x in titles])
            case_one = []
            for cell in rows_data[k]:
                case_one.append(cell.value)
            case_obj = Cases(*case_one)
            case_all.append(case_obj)
        return case_all

    def r_data_from_colunm(self, list1):
        """
        按list中行读取，表单所有该行数据
        每个用例存储在一个对象中
        :param list1: list1  -->   要求读取的列编号列表，eg:[1,3,5]则会读取所有用例的第1,3,5列
        :return: 包含所有用例对象的列表list
        """
        titles = []
        case_all = []
        for row in range(1, self.sheet.max_row + 1):
            case_one = []
            if row == 1:
                for column in list1:
                    titles.append(self.sheet.cell(row, column).value)
            else:
                for column in list1:
                    case_one.append(self.sheet.cell(row, column).value)
                case_all.append(dict(zip(titles, case_one)))
        return case_all

    def r_data_obj_from_column(self, list1):
        """
        按list中行读取，表单所有该行数据
        每个用例存储在一个对象中
        :param list1: list1  -->   要求读取的列编号列表，eg:[1,3,5]则会读取所有用例的第1,3,5列
        :return: 包含所有用例对象的列表list
        """
        titles = []
        case_all = []
        for row in range(1, self.sheet.max_row + 1):
            case_one = []
            if row == 1:
                for column in list1:
                    titles.append(self.sheet.cell(row, column).value)
            else:
                for column in list1:
                    case_one.append(self.sheet.cell(row, column).value)
                cases = namedtuple('Cases', titles)
                case_obj = cases(*case_one)
                case_all.append(case_obj)
        return case_all

    def w_data(self, row, column, data):
        self.sheet.cell(row, column, data)
        self.wb.save(self.filename)


if __name__ == '__main__':
    r = ReadExcel('cases.xlsx', 'Sheet1')
    print('---------------------------------------------------')
    data = r.r_data_from_colunm([1, 2, 3])
    print(data)

    print('---------------------------------------------------')
    data = r.r_data_obj_from_column([1, 2, 3])
    print(data)
