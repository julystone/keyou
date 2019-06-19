# encoding:utf8

# -----------------------------
# @File   :   excel_pc.py
# @Author :   July401
# @Date   :   2019/6/19
# @Email  :   july401@qq.com
# -----------------------------

import openpyxl

f = openpyxl.load_workbook('test1.xlsx')
sheet = f['Sheet1']

max_row = sheet.max_row
max_column = sheet.max_column
print(max_row, max_column)

for row in list(sheet.rows)[::-1]:
    if row[0].value is None:
        sheet.delete_rows(row[0].row, 1)

for column in list(sheet.columns)[::-1]:
    if column[0].value is None:
        sheet.delete_cols(column[0].column, 1)
f.save('test1.xlsx')
