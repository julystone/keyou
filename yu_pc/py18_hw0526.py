# -*-coding:utf8-*-
'''
============================
author:guxiangyu
time:2019/6/9
E-mail:317096158@qq.com
============================
'''
import openpyxl


class Case(object):
    def __init__(self, attrs):
        for item in attrs:
            setattr(self, item[0], item[1])


class ReadExcel(object):

    def __init__(self, file_name, sheet_name):
        self.wb = openpyxl.load_workbook(file_name)
        self.sheet = self.wb[sheet_name]

    # 读取excel数据

    def read_data_line(self, list1):
        max_r = self.sheet.max_row
        titles = []
        cases = []
        for row in range(1, max_r + 1):
            if row != 1:
                case_data = []
                for column in list1:
                    info = self.sheet.cell(row, column).value
                    case_data.append(info)
                case = dict(list(zip(titles, case_data)))
                cases.append(case)

            else:
                for column in list1:
                    title = self.sheet.cell(row, column).value
                    titles.append(title)
        return cases

    def read_data_obj(self, list1):
        max_r = self.sheet.max_row
        titles = []
        cases = []
        for row in range(1, max_r + 1):
            if row != 1:
                case_data = []
                for column in list1:
                    info = self.sheet.cell(row, column).value
                    case_data.append(info)
                case = dict(list(zip(titles, case_data)))
                case_obj = Case(case)
                cases.append(case_obj)

            else:
                for column in list1:
                    title = self.sheet.cell(row, column).value
                    titles.append(title)
        return cases


if __name__ == '__main__':
    r = ReadExcel('cases.xlsx', 'Sheet')
    data = r.read_data_line([1, 2, 3])
    data2 = r.read_data_obj([1, 3, 5])
    print(data)
    print(data2)
